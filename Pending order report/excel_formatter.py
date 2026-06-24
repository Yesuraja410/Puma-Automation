# -*- coding: utf-8 -*-
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Border styles
thin_side = Side(style='thin', color='A0A0A0')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Colors matching user requirements & reference image
COLOR_RED = 'FF0000'         # Overdue (SLA breached)
COLOR_ORANGE = 'FFC000'      # Handover Today (Today SLA)
COLOR_DARK_RED = 'C00000'    # Order status at NEW
COLOR_GREEN = '92D050'       # Within SLA (Future)
COLOR_LIGHT_GREEN = 'E2EFDA'  # Not reflected in OMS

FILL_RED = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')
FILL_ORANGE = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type='solid')
FILL_DARK_RED = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type='solid')
FILL_GREEN = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color=COLOR_LIGHT_GREEN, end_color=COLOR_LIGHT_GREEN, fill_type='solid')

FONT_WHITE_BOLD = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
FONT_NORMAL = Font(name='Calibri', size=11, bold=False)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)
FONT_TITLE = Font(name='Calibri', size=14, bold=True)

def _is_blank(val):
    s = str(val).strip()
    return not s or s.lower() in ("nan", "none", "nat", "null")

def get_date_status(date_str, ref_date_str):
    """
    Compare date_str to ref_date_str (both in 'DD-MM-YYYY' format).
    Returns 'Breached', 'Today', 'Future', or 'Unknown'.
    """
    if _is_blank(date_str) or _is_blank(ref_date_str):
        return 'Unknown'
    try:
        dt = datetime.strptime(date_str.strip(), '%d-%m-%Y')
        ref_dt = datetime.strptime(ref_date_str.strip(), '%d-%m-%Y')
        if dt < ref_dt:
            return 'Breached'
        elif dt == ref_dt:
            return 'Today'
        else:
            return 'Future'
    except Exception:
        return 'Unknown'

def autofit_columns(ws, min_width=10, padding=3):
    """Auto-adjusts columns width according to cells content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Determine maximum content length
        for cell in col:
            val = str(cell.value or '')
            # Handle linebreaks or formatting extensions
            if '\n' in val:
                lines_len = [len(l) for l in val.split('\n')]
                cell_len = max(lines_len) if lines_len else 0
            else:
                cell_len = len(val)
            if cell_len > max_len:
                max_len = cell_len
                
        ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

def format_data_sheet(ws, df):
    """Applies basic styling, bold headers, thin borders, and center alignment to all data cells."""
    ws.sheet_view.showGridLines = True
    
    # Header Row formatting
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = FONT_BOLD
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # Data Rows formatting (all showing center alignment as requested)
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
                
    autofit_columns(ws)

def add_country_sheets_to_workbook(wb, country, raw_df, pivot_df, summary_df, ref_date_str):
    """
    Adds country styled Summary and Data sheets to an existing workbook.
    """
    ws_summary = wb.create_sheet(title=f"{country} Summary")
    ws_summary.sheet_view.showGridLines = True
    
    # ── Title Block merged to match the pivot table width ──
    title_end_col = len(pivot_df.columns) if not pivot_df.empty else 6
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=title_end_col)
    
    title_cell = ws_summary.cell(row=2, column=1, value=f"Pending Orders - PUMA {country}")
    title_cell.font = FONT_TITLE
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply borders around Title cell block
    for col_idx in range(1, title_end_col + 1):
        ws_summary.cell(row=2, column=col_idx).border = thin_border
        
    # ── Pivot Table (Starting at A4) ──
    if not pivot_df.empty:
        # Write headers
        headers = list(pivot_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=4, column=col_idx, value=header)
            cell.font = FONT_BOLD
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Write rows
        # Classic Pivot: consecutive duplicate channels are left blank
        last_channel = None
        for row_idx, row_data in enumerate(pivot_df.itertuples(index=False), start=5):
            is_grand_total_row = (row_data[0] == "Grand Total")
            
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # Check for Channel column (1st column)
                if col_idx == 1:
                    if is_grand_total_row:
                        cell.value = "Grand Total"
                        cell.font = FONT_BOLD
                    else:
                        chan_val = str(val or '').strip()
                        if chan_val == last_channel:
                            cell.value = ""
                        else:
                            cell.value = chan_val
                            last_channel = chan_val
                            cell.font = FONT_BOLD
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Check for OMS Status column (2nd column)
                elif col_idx == 2:
                    if is_grand_total_row:
                        cell.value = ""
                    else:
                        cell.value = str(val or '').strip()
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = FONT_BOLD if is_grand_total_row else FONT_NORMAL
                # Check for Date Columns & Grand Total value
                else:
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Highlight counts > 0 based on date column SLA status
                    col_header = headers[col_idx - 1]
                    is_grand_total_col = (col_header == "Grand Total")
                    
                    if is_grand_total_row or is_grand_total_col:
                        cell.font = FONT_BOLD
                    else:
                        cell.font = FONT_NORMAL
                        # Highlight cell counts > 0 in data cells
                        if isinstance(val, (int, float)) and val > 0:
                            # Determine date status
                            status = get_date_status(col_header, ref_date_str)
                            if status == 'Today':
                                cell.fill = FILL_ORANGE
                            elif status == 'Future':
                                cell.fill = FILL_GREEN
                            elif status == 'Breached':
                                cell.fill = FILL_RED
                                cell.font = FONT_WHITE_BOLD
                                
    # ── Highlight Metrics Table (Starting after the Pivot Table with 1 column gap) ──
    start_col = (len(pivot_df.columns) + 2) if not pivot_df.empty else 11
    metrics_map = {}
    if not summary_df.empty:
        metrics_map = summary_df.set_index("Metric")["Count"].to_dict()
        
    metrics_list = [
        ("Overdue (SLA breached)", "Overdue", FILL_RED, FONT_WHITE_BOLD),
        ("Handover today (Today SLA)", "Handover Today", FILL_ORANGE, FONT_BLACK_BOLD),
        ("Order Status at New", "Order status at NEW", FILL_DARK_RED, FONT_WHITE_BOLD),
        ("Within SLA (Future)", "Within SLA", FILL_GREEN, FONT_BLACK_BOLD),
        ("Not reflecting in OM", "Not reflected in OMS", FILL_LIGHT_GREEN, FONT_BLACK_BOLD)
    ]
    
    for idx, (original_name, display_name, fill, font) in enumerate(metrics_list):
        row_pos = 4 + idx
        
        # Metric Label
        cell_lbl = ws_summary.cell(row=row_pos, column=start_col, value=display_name)
        cell_lbl.fill = fill
        cell_lbl.font = font
        cell_lbl.alignment = Alignment(horizontal='center', vertical='center')
        cell_lbl.border = thin_border
        
        # Metric Value
        val = metrics_map.get(original_name, 0)
        if display_name == "Not reflected in OMS" and (val == 0 or val == "-"):
            val = "-"
            
        cell_val = ws_summary.cell(row=row_pos, column=start_col + 1, value=val)
        cell_val.font = FONT_BOLD
        cell_val.alignment = Alignment(horizontal='center', vertical='center')
        cell_val.border = thin_border
        
    autofit_columns(ws_summary)
    
    # == 2. Data Sheet =========================================================
    ws_data = wb.create_sheet(title=f"{country} Data")
    
    # Write Header
    if not raw_df.empty:
        cols = list(raw_df.columns)
        for col_idx, col_name in enumerate(cols, start=1):
            ws_data.cell(row=1, column=col_idx, value=col_name)
            
        # Write rows
        for row_idx, row_data in enumerate(raw_df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws_data.cell(row=row_idx, column=col_idx, value=val)
                
        format_data_sheet(ws_data, raw_df)

def generate_excel_workbook(country, raw_df, pivot_df, summary_df, ref_date_str):
    """
    Creates a country-specific workbook (Summary + Data).
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    add_country_sheets_to_workbook(wb, country, raw_df, pivot_df, summary_df, ref_date_str)
    
    # Rename f"{country} Summary" -> "Summary" and f"{country} Data" -> "Data"
    wb.worksheets[0].title = "Summary"
    wb.worksheets[1].title = "Data"
    
    return wb
